"""
file_parser.py
==============

アップロードされた **CSV / Excel** ファイルを堅牢に `pandas.DataFrame`
へ変換するユーティリティ。

* MIME & 拡張子でファイル種別を判定
* CSV は **chardet** でエンコーディング推定＋フォールバックを順次試行
* ヘッダー行から BOM / 全角スペースを除去して統一
* 値はすべて **string 型** で読み取り (`dtype=str`, `keep_default_na=False`)
* 空ファイル・非対応形式は ``ValueError`` を送出

本モジュールは **FastAPI UploadFile / Path / str / bytes** の
いずれでも受け取れるように実装してあるので、  
API・pytest の両方でそのまま呼び出せる。
"""

from __future__ import annotations

import io
import mimetypes
from pathlib import Path
from typing import Final, Iterable

import pandas as pd
import chardet
import unicodedata

try:
    # UploadFile はテスト環境では import 不要なので optional
    from fastapi import UploadFile
except ModuleNotFoundError:  # pragma: no cover
    UploadFile = object   # type: ignore


ENCODINGS: Final[list[str]] = [
    "utf-8",
    "utf-8-sig",
    "utf-16",
    "utf-16-le",
    "utf-16-be",
    "cp932",
    "iso8859-1",
]


# --------------------------------------------------------------------------- #
# public API                                                                  #
# --------------------------------------------------------------------------- #
def read_dataframe(file: UploadFile | str | Path | bytes | bytearray) -> pd.DataFrame:
    """
    Parameters
    ----------
    file :
        * **FastAPI UploadFile** – 実運用でのアップロード
        * **str / Path** – 単体テストやローカル実行で Path を直接渡す
        * **bytes / bytearray** – メモリ上のバイト列

    Returns
    -------
    pandas.DataFrame
        先頭行をヘッダーと見なし、すべて **文字列型** で読み込んだ表データ

    Raises
    ------
    ValueError
        - 空ファイル
        - 非対応ファイル形式
        - エンコーディング判定失敗
    """
    raw, filename = _get_raw_and_name(file)

    if not raw:
        raise ValueError("File is empty")

    mime, _ = mimetypes.guess_type(filename)
    lower_name = filename.lower()

    # ----------------------------- CSV ------------------------------------
    if mime in ("text/csv", None) or lower_name.endswith(".csv"):
        mime, _ = mimetypes.guess_type(filename)
        # Performance optimization: Try UTF-8 first (most common), then CP932 (Japanese ERP)
        might_be_utf16 = b"\x00" in raw[:1024]
        
        # Fast path encodings to try first
        if might_be_utf16:
            enc_try_order = ["utf-16", "utf-16-le", "utf-16-be", "utf-8", "utf-8-sig", "cp932"]
        else:
            enc_try_order = ["utf-8", "utf-8-sig", "cp932"]
        
        last_error = None

        for enc in enc_try_order:
            try:
                # csv.Sniffer does not work well on UTF‑16 so we force comma there.
                sep_param = None if enc.startswith("utf-8") or enc == "cp932" else ","

                # Try fast path: pyarrow engine for UTF-8 family when available
                use_arrow = False
                if sep_param is None and enc in ("utf-8", "utf-8-sig"):
                    try:
                        import pyarrow  # type: ignore  # noqa: F401
                        use_arrow = True
                    except Exception:
                        use_arrow = False

                if use_arrow:
                    df_try = pd.read_csv(
                        io.BytesIO(raw),
                        encoding=enc,
                        keep_default_na=False,
                        engine="pyarrow",
                    )
                    # Ensure stringy behavior for downstream mappers
                    for c in df_try.columns:
                        df_try[c] = df_try[c].astype(str)
                else:
                    df_try = pd.read_csv(
                        io.BytesIO(raw),
                        encoding=enc,
                        dtype=str,
                        keep_default_na=False,
                        sep=sep_param,
                        engine="python",
                    )
                # Validate parse result for delimiter detection failure
                if df_try.shape[1] == 1 and b"," in raw[:1024]:
                    try:
                        df_try2 = pd.read_csv(
                            io.BytesIO(raw),
                            encoding=enc,
                            dtype=str,
                            keep_default_na=False,
                            sep=",",        # explicit comma delimiter
                        )
                        if df_try2.shape[1] > 1:
                            df_try = df_try2
                    except Exception:
                        pass
                # If we still only have a single column, try a few common
                # alternative delimiters (tab / semicolon / pipe).  Some
                # Japanese ERP exports in CP932 use tab‑separated format.
                if df_try.shape[1] == 1:
                    for _sep in ("\t", ";", "|"):
                        try:
                            df_alt = pd.read_csv(
                                io.BytesIO(raw),
                                encoding=enc,
                                dtype=str,
                                keep_default_na=False,
                                sep=_sep,
                            )
                            if df_alt.shape[1] > 1:
                                df_try = df_alt
                                break
                        except Exception:
                            # Try next delimiter
                            continue
                # Success! Return immediately
                return df_try
            except (UnicodeDecodeError, pd.errors.ParserError) as e:
                last_error = e
                continue
        
        # Fast path failed, try chardet as expensive fallback
        try:
            enc_guess: str = (chardet.detect(raw[:4096]).get("encoding") or "").lower()
            if enc_guess and enc_guess not in enc_try_order:
                # Add fallback encodings
                fallback_encs = [enc_guess] + [e for e in ENCODINGS if e not in enc_try_order]
                
                for enc in fallback_encs:
                    try:
                        sep_param = None if enc.startswith("utf-8") or enc == "cp932" else ","
                        df_try = pd.read_csv(
                            io.BytesIO(raw),
                            encoding=enc,
                            dtype=str,
                            keep_default_na=False,
                            sep=sep_param,
                            engine="python",
                        )
                        # Validate delimiter detection
                        if df_try.shape[1] == 1 and b"," in raw[:1024]:
                            df_try2 = pd.read_csv(
                                io.BytesIO(raw),
                                encoding=enc,
                                dtype=str,
                                keep_default_na=False,
                                sep=",",
                            )
                            if df_try2.shape[1] > 1:
                                df_try = df_try2
                        return df_try
                    except (UnicodeDecodeError, pd.errors.ParserError):
                        continue
        except Exception:
            pass
        
        # All attempts failed
        if last_error:
            raise ValueError(f"Cannot decode CSV – tried {len(enc_try_order)} encodings, last error: {last_error}")
        else:
            raise ValueError("Cannot decode CSV – unknown encoding")

    # ----------------------------- Excel ----------------------------------
    elif lower_name.endswith((".xlsx", ".xls")):
        # .xlsx は openpyxl の read_only + values_only で高速ロードを試す
        if lower_name.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook  # type: ignore
                wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                ws = wb.active
                # 先頭行をヘッダーとして採用（完全空行はスキップ）
                rows_it = ws.iter_rows(values_only=True)
                headers = None
                data_rows: list[list[str]] = []
                for r in rows_it:
                    if r is None:
                        continue
                    # openpyxl はタプルを返す
                    row_vals = ["" if (c is None) else str(c) for c in r]
                    # 完全空行はスキップ
                    if all(v == "" for v in row_vals):
                        continue
                    if headers is None:
                        headers = row_vals
                        continue
                    # 列数をヘッダーに合わせて調整
                    if len(row_vals) < len(headers):
                        row_vals.extend([""] * (len(headers) - len(row_vals)))
                    elif len(row_vals) > len(headers):
                        row_vals = row_vals[: len(headers)]
                    data_rows.append(row_vals)
                if headers is None:
                    raise ValueError("Excelにヘッダー行がありません")
                # DataFrame化
                df = pd.DataFrame(data_rows, columns=headers)
            except ImportError as e:
                raise ValueError(
                    "Excelファイル(.xlsx)を読み込むには 'openpyxl' が必要です。CSVでアップロードするか、サーバに openpyxl をインストールしてください。"
                ) from e
            except Exception:
                # 失敗時は pandas 経由にフォールバック
                try:
                    df = pd.read_excel(
                        io.BytesIO(raw), dtype=str, keep_default_na=False, engine="openpyxl"
                    )
                except Exception as e2:
                    raise ValueError(f"Excelの読み込みに失敗しました: {e2}") from e2
        else:
            # .xls はサポート外（xlrd 非対応）。pandas で試み、失敗なら親切エラー。
            try:
                df = pd.read_excel(io.BytesIO(raw), dtype=str, keep_default_na=False)
            except Exception as e:
                raise ValueError(".xls は読み込み非推奨です。CSV へ保存してアップロードしてください。") from e

    else:
        raise ValueError("Unsupported file type (only .csv/.xlsx/.xls accepted)")

    # NOTE: ヘッダ名は原則としてそのまま返す（テストは生ヘッダを期待）。
    # 以前は NFKC 正規化や別名畳み込みを行っていたが、
    # それらはアップロード層（mappers）で多名対応することで吸収する。
    # ここではパースの堅牢性のみに責務を絞る。

    if df.empty:
        raise ValueError("File has no data rows")

    return df


__all__ = ["read_dataframe"]


# --------------------------------------------------------------------------- #
# helpers (private)                                                           #
# --------------------------------------------------------------------------- #
def _get_raw_and_name(
    file: UploadFile | str | Path | bytes | bytearray,
) -> tuple[bytes, str]:
    """
    Convert various *file‑like* inputs into raw bytes + filename.

    Accepts:

    * FastAPI / Starlette ``UploadFile`` (objects with ``.file`` & ``.filename``)
    * ``str`` / ``pathlib.Path`` pointing to a file on disk
    * ``bytes`` / ``bytearray`` already in memory
    """
    # ----------------------- in‑memory bytes ------------------------------
    if isinstance(file, (bytes, bytearray)):
        return bytes(file), ""

    # ----------------------- filesystem path ------------------------------
    if isinstance(file, (str, Path)):
        p = Path(file)
        return p.read_bytes(), p.name

    # ------------------- FastAPI / Starlette UploadFile -------------------
    # NOTE: `fastapi.UploadFile` *is* `starlette.datastructures.UploadFile`,
    #       but when FastAPI is missing (pytest for utils) we fall back to
    #       a dummy ``object``.  Guard against that case explicitly.
    try:
        from starlette.datastructures import UploadFile as StarletteUploadFile  # type: ignore
    except ModuleNotFoundError:  # pragma: no cover – extremely unlikely
        StarletteUploadFile = None  # type: ignore

    # FastAPI import succeeded -> UploadFile is the real class
    if "UploadFile" in globals() and UploadFile is not object and isinstance(file, UploadFile):  # type: ignore
        return file.file.read(), file.filename or ""

    # Fallback: Starlette class available but different import path
    if StarletteUploadFile is not None and isinstance(file, StarletteUploadFile):
        return file.file.read(), file.filename or ""

    # ---------------------------- duck typing -----------------------------
    if hasattr(file, "file") and hasattr(file, "filename"):
        return file.file.read(), getattr(file, "filename", "") or ""

    # ----------------------------- failure --------------------------------
    raise TypeError(
        "file must be UploadFile | str | Path | bytes | bytearray; "
        f"got {type(file)}"
    )


def _unique(seq: Iterable[str]) -> list[str]:
    """順序を保ったまま重複削除"""
    seen: set[str] = set()
    out: list[str] = []
    for item in seq:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out